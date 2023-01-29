import openpyxl as xl
from openpyxl.chart import BarChart, Reference


# def name_workbook(filename):
#     wb = xl.load_workbook(filename)
#     sheet = wb['Sheet1']
#     return sheet, wb
#
#
# # dOne = name_workbook.cell(1, 4)
# # dOne.value = 'corrected_price'
#
#
# def correct_prices(sheet):
#     for row in range(2, sheet.max_row + 1):
#         cell = sheet.cell(row, 3)
#         corrected_price = cell.value * 0.9
#         corrected_price_cell = sheet.cell(row, 4)
#         corrected_price_cell.value = corrected_price
#
#
# def create_chart(sheet):
#     values = Reference(sheet,
#               min_row=2,
#               max_row=sheet.max_row,
#               min_col=4,
#               max_col=4)
#
#     chart = BarChart()
#     chart.add_data(values)
#     sheet.add_chart(chart, 'e2')
#
#
# wb.save('transactions2.xlsx')


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    dOne = sheet.cell(1, 4)
    dOne.value = 'corrected_price'

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

